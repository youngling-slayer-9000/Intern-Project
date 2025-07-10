import os
import django
from datetime import datetime, timedelta
from openpyxl import load_workbook
import sys

# Setup Django
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'AttendanceSystem.settings')
django.setup()

from home.models import AttendanceRecord

def load_data():
    file_path = "emloyeeAttendanceUpdated.xlsx"  # ✅ updated file name
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            
            emp_id, name, dept, date_str, _, first_punch_str, last_punch_str, total_time_str, _, _, branch = row

            if not name:
                print(f"⚠️ Skipped row due to missing name: {row}")
                continue



            date = datetime.strptime(str(date_str), "%d-%m-%Y").date()
            first_punch = datetime.strptime(first_punch_str, "%H:%M").time()
            last_punch = datetime.strptime(last_punch_str, "%H:%M").time()

            hours, minutes = map(int, total_time_str.split(":"))
            total_time = timedelta(hours=hours, minutes=minutes)

            AttendanceRecord.objects.create(
                employee_id=str(emp_id).zfill(4),
                first_name=name,
                department=dept,
                branch=branch,
                date=date,
                first_punch=first_punch,
                last_punch=last_punch,
                total_time=total_time
            )
        except Exception as e:
            print("❌ Error on row:", row, "| Error:", e)

if __name__ == "__main__":
    load_data()
    print("✅ Attendance data loaded successfully from employeeAttendanceUpdated.xlsx.")
